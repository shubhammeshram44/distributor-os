import csv
import os
import uuid
import logging
import io
from typing import List, Dict, Optional, Any
from sqlalchemy import select, func
from sqlalchemy.orm import Session
from pydantic import BaseModel
import google.generativeai as genai
from app.models.ingestion import IngestionJob, IngestionStaging
from app.models.product import Product, ProductAlias
from app.models.inventory import Inventory
from app.config import settings
from app.database import tenant_context, with_db_retry
import openpyxl

# Event Discriminator & Ingestion Imports
from datetime import datetime
from app.models.tenant import DistributorTenant
from app.models.customer import Customer, CustomerAlias
from app.models.order import Order, OrderLineItem, OrderStateLedger
from app.services.gemini_service import GeminiService
from app.utils.phone import normalize_phone_number, get_phone_number_variants
from app.services.tenant_service import DEMO_TENANT_ID

logger = logging.getLogger(__name__)

# ── PRE-FILTER: Skip Gemini for obvious non-order messages ──────────────────
# Saves ~40% of Gemini API calls at no cost to order accuracy.
# Only skip if message is SHORT and matches known non-order patterns.

NON_ORDER_PHRASES = {
    # Acknowledgements
    "ok", "okay", "ok bhaiya", "okay bhaiya", "theek hai", "theek",
    "haan", "haan bhaiya", "ha", "han", "haa",
    # Received confirmations  
    "received", "mil gaya", "mil gayi", "aa gaya", "aa gayi",
    "mila", "pakad liya", "le liya",
    # Thanks
    "thanks", "thank you", "shukriya", "dhanyawad",
    # Simple yes/no
    "yes", "no", "nahi", "nhi", "na",
    # Emojis only messages (common reactions)
    "👍", "👍🏻", "👍🏼", "👍🏽", "✅", "🙏", "🙏🏻", "❤️", "😊",
    # Common greetings (not orders)
    "hello", "hi", "helo", "namaste", "namaskar",
    # Other common responses
    "kal bata ta", "kal batata", "baad mein", "thoda ruko",
    "wait", "ruko", "ek minute", "1 minute",
}

def is_obvious_non_order(text: str) -> bool:
    """
    Returns True if message is clearly not an order.
    Only skips very short messages that exactly match known non-order phrases.
    Never skips messages longer than 30 characters — safety margin.
    """
    if not text:
        return True
    
    # Never skip longer messages — could be an order with extra context
    if len(text.strip()) > 30:
        return False
    
    normalized = text.strip().lower()
    
    # Exact match
    if normalized in NON_ORDER_PHRASES:
        return True
    
    # Message is only emojis/special chars (no alphanumeric content)
    import re
    if not re.search(r'[a-zA-Z0-9\u0900-\u097F]', normalized):
        return True
    
    return False


class HeaderMapping(BaseModel):
    SKU: str
    Quantity: str
    Quantity_Committed: str

class IngestionService:
    # Class-level cache to share parsed phone mappings across instances and reduce database load
    _distributor_phone_cache: Dict[uuid.UUID, str] = {}

    @classmethod
    def invalidate_tenant_cache(cls, tenant_id: uuid.UUID) -> None:
        """
        Invalidates the cached business number for the given tenant ID.
        Should be called when distributor settings are updated.
        """
        if tenant_id in cls._distributor_phone_cache:
            cls._distributor_phone_cache.pop(tenant_id, None)
            logger.info("IngestionService: Invalidated cached distributor phone for tenant %s", tenant_id)

    def __init__(self):
        self.api_key = settings.GEMINI_API_KEY
        self.enabled = bool(self.api_key)
        if self.enabled:
            try:
                genai.configure(api_key=self.api_key)
                self.model = genai.GenerativeModel("gemini-1.5-flash")
            except Exception as e:
                logger.error(f"Failed to initialize Gemini in Ingestion: {e}")
                self.enabled = False

    def parse_file(self, file_content: bytes, filename: str) -> List[Dict[str, Any]]:
        """
        Parses CSV or Excel bytes into a list of row dictionaries.
        """
        rows = []
        if filename.endswith(".csv"):
            text_stream = io.StringIO(file_content.decode("utf-8"))
            reader = csv.DictReader(text_stream)
            for row in reader:
                rows.append(dict(row))
        elif filename.endswith((".xlsx", ".xls")):
            wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row_cells in ws.iter_rows(min_row=2, values_only=True):
                if any(row_cells): # Skip fully empty rows
                    row_dict = {headers[i]: row_cells[i] for i in range(len(headers)) if i < len(row_cells)}
                    rows.append(row_dict)
        else:
            raise ValueError("Unsupported file format. Please upload CSV or Excel.")
        return rows

    def get_column_mapping(self, headers: List[str]) -> Dict[str, str]:
        """
        Maps raw file headers to canonical keys: 'SKU', 'Quantity', 'Quantity_Committed'.
        """
        if self.enabled:
            try:
                prompt = (
                    "Analyze the following list of uploaded file headers. "
                    "Map these headers to our canonical fields: 'SKU', 'Quantity', 'Quantity_Committed'. "
                    "Return a JSON object where the keys are our canonical fields and the values are "
                    "the exact matched headers from the uploaded file list. "
                    f"Uploaded Headers: {headers}"
                )

                generation_config = {
                    "response_mime_type": "application/json",
                    "response_schema": HeaderMapping,
                }

                response = self.model.generate_content(
                    contents=[prompt],
                    generation_config=generation_config
                )
                import json
                mapping = json.loads(response.text)
                return mapping
            except Exception as e:
                logger.error(f"Gemini column mapping failed: {e}. Falling back to rule-based mapper.")

        return self._fallback_header_mapper(headers)

    def _fallback_header_mapper(self, headers: List[str]) -> Dict[str, str]:
        """
        Regex/keyword header mapper for fallback testing.
        """
        mapping = {"SKU": "", "Quantity": "", "Quantity_Committed": ""}
        normalized_headers = {h.lower().replace("_", "").replace(" ", ""): h for h in headers if h}

        # SKU matching
        for k in ["sku", "skuid", "itemcode", "productcode", "itemname", "product", "item"]:
            if k in normalized_headers:
                mapping["SKU"] = normalized_headers[k]
                break

        # Quantity matching
        for k in ["qty", "quantity", "stock", "quantityonhand", "onhand", "stockqty"]:
            if k in normalized_headers:
                mapping["Quantity"] = normalized_headers[k]
                break

        # Quantity Committed matching
        for k in ["committed", "quantitycommitted", "reserved", "committedqty", "allocated"]:
            if k in normalized_headers:
                mapping["Quantity_Committed"] = normalized_headers[k]
                break

        # Fallback defaults if no match found
        if not mapping["SKU"] and headers:
            mapping["SKU"] = headers[0]
        if not mapping["Quantity"] and len(headers) > 1:
            mapping["Quantity"] = headers[1]
        if not mapping["Quantity_Committed"] and len(headers) > 2:
            mapping["Quantity_Committed"] = headers[2]

        return mapping

    @with_db_retry
    def ingest_file(
        self,
        db: Session,
        tenant_id: uuid.UUID,
        file_content: bytes,
        filename: str
    ) -> IngestionJob:
        """
        Main entry point to upload a file, write to IngestionStaging, and validate row-by-row.
        """
        tenant_context.set(tenant_id)

        # 1. Parse File
        try:
            raw_rows = self.parse_file(file_content, filename)
        except Exception as e:
            logger.error(f"File parsing error: {e}")
            job = IngestionJob(
                tenant_id=tenant_id,
                source="Excel" if filename.endswith((".xlsx", ".xls")) else "CSV",
                status="Failed",
                filename=filename,
                total_rows=0
            )
            db.add(job)
            db.commit()
            return job

        # 2. Get Headers and Columns mapping
        headers = list(raw_rows[0].keys()) if raw_rows else []
        col_map = self.get_column_mapping(headers)

        # 3. Create Ingestion Job
        job = IngestionJob(
            tenant_id=tenant_id,
            source="Excel" if filename.endswith((".xlsx", ".xls")) else "CSV",
            status="Processing",
            filename=filename,
            total_rows=len(raw_rows),
            successful_rows=0,
            failed_rows=0
        )
        db.add(job)
        db.flush()

        # 4. Insert into Staging
        staging_records = []
        for row in raw_rows:
            staging_rec = IngestionStaging(
                tenant_id=tenant_id,
                job_id=job.id,
                raw_data=row,
                status="Staged"
            )
            db.add(staging_rec)
            staging_records.append(staging_rec)
        db.flush()

        # 5. Process Row-by-Row
        sku_col = col_map.get("SKU")
        qty_col = col_map.get("Quantity")
        committed_col = col_map.get("Quantity_Committed")

        for staging in staging_records:
            row_data = staging.raw_data
            
            # Start nested transaction (Savepoint) for row-level rollback isolation
            nested = db.begin_nested()
            try:
                # Extract fields
                raw_sku = str(row_data.get(sku_col) or "").strip()
                raw_qty = row_data.get(qty_col)
                raw_committed = row_data.get(committed_col)

                if not raw_sku:
                    raise ValueError("SKU field is empty in this row.")

                # Convert quantities to integers safely
                try:
                    qty = int(float(raw_qty)) if raw_qty is not None else 0
                    committed = int(float(raw_committed)) if raw_committed is not None else 0
                except (ValueError, TypeError):
                    raise ValueError(f"Invalid quantity formats. Qty: {raw_qty}, Committed: {raw_committed}")

                if qty < 0 or committed < 0:
                    raise ValueError("Quantities cannot be negative.")

                # Check if SKU matches a Product
                product_query = select(Product).where(Product.sku_id == raw_sku)
                product = db.execute(product_query).scalar_one_or_none()

                if not product:
                    # Try matching via ProductAlias
                    product_alias_query = (
                        select(Product)
                        .join(Product.aliases)
                        .where(func.lower(ProductAlias.alias_name) == raw_sku.lower())
                    )
                    product = db.execute(product_alias_query).scalar_one_or_none()

                if not product:
                    raise ValueError(f"SKU or Alias '{raw_sku}' does not exist in product catalog.")

                # Check or Create Inventory Record
                inventory_query = select(Inventory).where(Inventory.sku_id == product.id)
                inventory = db.execute(inventory_query).scalar_one_or_none()

                if not inventory:
                    inventory = Inventory(
                        tenant_id=tenant_id,
                        sku_id=product.id,
                        location="Staging Area", # Default
                        quantity_on_hand=qty,
                        quantity_committed=committed,
                        low_stock_threshold=10
                    )
                    db.add(inventory)
                else:
                    inventory.quantity_on_hand = qty
                    inventory.quantity_committed = committed

                # Save Staging State
                staging.status = "Validated"
                job.successful_rows += 1
                
                # Commit nested transaction
                nested.commit()

            except Exception as row_error:
                nested.rollback() # Rollback only changes inside this row's savepoint
                
                # Update staging logs in parent transaction context
                staging.status = "Failed"
                staging.error_message = str(row_error)
                job.failed_rows += 1
                logger.warning(f"File ingestion row failed: {row_error}")

            # Flush individual changes to parent db transaction
            db.flush()

        job.status = "Completed"
        db.commit()
        return job

    @with_db_retry
    def ingest_message(
        self,
        db: Session,
        tenant_id: uuid.UUID,
        sender_phone: str,
        message_text: str,
        sender_jid: Optional[str] = None,
        from_me: bool = False,
    ) -> dict:
        """
        5-Layer Ingestion Triage Pipeline for incoming WhatsApp orders.

        Layer 1 – JID Validation      : Drop non-customer JIDs (groups, broadcasts).
        Layer 2 – Sender Isolation    : Drop messages sent from the distributor's own number.
        Layer 3 – Length Filter       : Drop trivially short texts (< 5 chars).
        Layer 4 – Gemini Flash Intent : Drop non-order messages via lightweight LLM classification.
        Layer 5 – Customer Whitelist  : Drop senders not registered as customers for this tenant.
        """
        tenant_context.set(tenant_id)

        # ── LAYER 1: JID Validation ─────────────────────────────────────────────
        # When a raw JID is available, ensure it belongs to an individual chat
        # (@s.whatsapp.net). Group chats (@g.us) and broadcast lists (@lid / @broadcast)
        # are dropped here before any further processing.
        if sender_jid:
            if not sender_jid.endswith("@s.whatsapp.net"):
                logger.info(
                    "IngestionService: Layer 1 – Dropping non-individual JID: %s", sender_jid
                )
                return {"status": "ignored", "reason": "invalid_jid_type"}

        # ── LAYER 2: Sender Isolation ────────────────────────────────────────────
        # Resolve cached distributor business number and drop messages that originated
        # from the distributor's own active WhatsApp link (fromMe == True or matching
        # normalized phone).
        if from_me:
            logger.info(
                "IngestionService: Layer 2 – Dropping distributor self-message (fromMe=True)"
            )
            return {"status": "ignored", "reason": "distributor_self_message"}

        distributor_phone = self._distributor_phone_cache.get(tenant_id)
        if not distributor_phone:
            tenant = db.query(DistributorTenant).filter(DistributorTenant.id == tenant_id).first()
            if tenant and tenant.whatsapp_order_phone:
                distributor_phone = tenant.whatsapp_order_phone
                self._distributor_phone_cache[tenant_id] = distributor_phone

        norm_sender = normalize_phone_number(sender_phone)
        norm_distributor = normalize_phone_number(distributor_phone) if distributor_phone else ""

        if norm_distributor and norm_sender == norm_distributor:
            logger.info(
                "IngestionService: Layer 2 – Dropping distributor business number %s.",
                norm_sender,
            )
            return {"status": "ignored", "reason": "distributor_self_message"}

        # ── LAYER 3: Length Filter ───────────────────────────────────────────────
        # Drop trivially short messages that can never represent a real order
        # (e.g. "Ok", "👍", "Haan", bare emoji replies).
        clean_text = message_text.strip()
        if len(clean_text) < 5:
            logger.info(
                "IngestionService: Layer 3 – Dropping short message (len=%d): '%s'",
                len(clean_text),
                clean_text,
            )
            return {"status": "ignored", "reason": "message_too_short"}

        # ── LAYER 4: Gemini Flash Intent Gatekeeper ──────────────────────────────
        # A lightweight call to gemini-2.5-flash classifies whether the message
        # expresses business purchase intent. Non-order messages are dropped here
        # before opening a database write session.
        #
        # Safety: when GEMINI_API_KEY is absent (e.g. CI / test environments)
        # the gate is bypassed and the message is passed through automatically.
        class IntentCheck(BaseModel):
            is_order: bool

        if self.enabled:
            try:
                intent_model = genai.GenerativeModel(
                    model_name="gemini-2.5-flash",
                    system_instruction=(
                        "Classify if the inbound text expresses any business intent to buy, "
                        "restock, or inquire about purchasing stock items (even shorthand "
                        "entries like 'Liril 50'). "
                        "Return false strictly for simple greetings, casual confirmations "
                        "like 'thanks' or 'received', emojis, or statements completely "
                        "unrelated to ordering product inventory."
                    ),
                )
                intent_response = intent_model.generate_content(
                    contents=[clean_text],
                    generation_config={
                        "response_mime_type": "application/json",
                        "response_schema": IntentCheck,
                    },
                )
                import json as _json
                intent_data = _json.loads(intent_response.text)
                is_order: bool = intent_data.get("is_order", True)
            except Exception as intent_exc:
                # Fail open: if the gatekeeper call errors, allow the message through
                # so we never silently drop a real order due to an API outage.
                logger.warning(
                    "IngestionService: Layer 4 – Intent check failed (%s); passing through.",
                    str(intent_exc),
                )
                is_order = True

            if not is_order:
                logger.info(
                    "IngestionService: Layer 4 – Non-order message classified by Gemini Flash. "
                    "Skipping DB write. text='%s'",
                    clean_text,
                )
                # Best-effort: a non-order reply from a KNOWN customer might be a
                # payment promise ("I'll pay by Friday") — worth capturing even
                # though we're not creating an order for this message. Never
                # affects the "ignored" response either way.
                try:
                    if sender_phone:
                        phone_variants = get_phone_number_variants(sender_phone)
                        promise_customer = (
                            db.query(Customer).join(CustomerAlias)
                            .filter(CustomerAlias.alias_value.in_(phone_variants)).first()
                            or db.query(Customer).filter(Customer.phone_number.in_(phone_variants)).first()
                        )
                        if promise_customer:
                            from app.services.payment_promise_service import detect_and_record_promise
                            detect_and_record_promise(db, tenant_id, promise_customer, clean_text)
                except Exception as promise_exc:
                    logger.warning("IngestionService: Layer 4 – Payment promise detection failed silently: %s", str(promise_exc))
                return {"status": "ignored", "reason": "non_order_intent"}
        else:
            logger.debug(
                "IngestionService: Layer 4 – Gemini API key absent; bypassing intent gate."
            )

        # ── LAYER 5: Customer Whitelist ──────────────────────────────────────────
        # Only registered customers for this tenant may create orders.
        customer = None
        if sender_phone:
            phone_variants = get_phone_number_variants(sender_phone)
            logger.info("IngestionService: Generated phone variants for lookup: %s", phone_variants)
            customer = db.query(Customer).join(CustomerAlias).filter(CustomerAlias.alias_value.in_(phone_variants)).first()
            if not customer:
                customer = db.query(Customer).filter(Customer.phone_number.in_(phone_variants)).first()

        if not customer:
            logger.warning("IngestionService: Layer 5 – Sender %s is not whitelisted for tenant %s", norm_sender, tenant_id)
            return {
                "status": "ignored",
                "message": f"Unknown customer phone number (not whitelisted): {norm_sender}",
                "job_id": None,
                "successful_rows": 0,
                "failed_rows": 0,
                "error_message": None
            }

        # ── PRE-FILTER: Skip Gemini for obvious non-order messages ──────────────────
        # Saves ~40% of Gemini API calls at no cost to order accuracy.
        # Only skip if message is SHORT and matches known non-order patterns.
        if is_obvious_non_order(message_text):
            logger.info(
                "IngestionService: Pre-filter skipped Gemini for non-order message: '%s'",
                message_text[:50]
            )
            return {"status": "ignored", "reason": "non_order_intent"}

        # ── CORE INGESTION: Order Parsing & SKU Matching ─────────────────────────
        gemini_service = GeminiService()
        parsed_order = gemini_service.parse_order_text(message_text)
        logger.info("IngestionService: text='%s' parsed_result=%s", message_text, parsed_order.model_dump_json())

        raw_tokens = []
        if parsed_order.items:
            for item in parsed_order.items:
                raw_tokens.append({
                    "text_token": item.raw_product_name,
                    "qty": item.quantity
                })

        # Fallback logic if LLM parsing yields no tokens
        if not raw_tokens:
            raw_tokens.append({
                "text_token": message_text,
                "qty": 1
            })

        # Match tokens against database catalog
        parsed_items = []
        has_unmatched = False

        # Helper to get the best product from alias query results, avoiding MOCK products if possible
        def get_best_product_for_alias_query(query):
            matches = query.all()
            if not matches:
                return None
            for am in matches:
                p = db.query(Product).filter_by(id=am.product_id).first()
                if p and "MOCK" not in p.sku_id:
                    return p
            return db.query(Product).filter_by(id=matches[0].product_id).first()

        # Helper to get the best product from product query results, avoiding MOCK products if possible
        def get_best_product_for_prod_query(query):
            matches = query.all()
            if not matches:
                return None
            for p in matches:
                if "MOCK" not in p.sku_id:
                    return p
            return matches[0]

        # Load once before the token loop
        all_tenant_aliases = db.query(ProductAlias).filter(
            ProductAlias.tenant_id == tenant_id
        ).all()

        all_tenant_products = db.query(Product).filter(
            Product.tenant_id == tenant_id,
            Product.is_active == True
        ).all()

        for token_entry in raw_tokens:
            token = token_entry["text_token"]
            qty = token_entry["qty"]
            product = None
            matched_in_phase2 = False
            matched_in_phase3 = False

            # Phase 1: Exact Match (Case-insensitive) against active tenant's aliases and product SKUs
            product = get_best_product_for_alias_query(
                db.query(ProductAlias).filter(
                    ProductAlias.tenant_id == tenant_id,
                    ProductAlias.alias_name.ilike(token)
                )
            )
            if not product:
                product = get_best_product_for_prod_query(
                    db.query(Product).filter(
                        Product.tenant_id == tenant_id,
                        Product.sku_id.ilike(token)
                    )
                )

            # Phase 2: Word overlap scoring
            token_lower = token.lower()
            token_words = []
            if not product:
                import re

                # Extract meaningful words from token (ignore words ≤ 2 chars like "of", "and", "a")
                token_words = [w for w in re.split(r'[\s\-_\./]+', token_lower) 
                               if w and len(w) > 2]
                
                if token_words:
                    best_product = None
                    best_score = 0.0

                    # Build candidate pool: aliases + product searchable fields
                    # Each candidate maps a searchable string → Product
                    candidates: list[tuple[str, Product]] = []
                    
                    for alias in all_tenant_aliases:
                        p = next((prod for prod in all_tenant_products if prod.id == alias.product_id), None)
                        if not p:
                            p = db.query(Product).filter_by(id=alias.product_id).first()
                        if p and "MOCK" not in p.sku_id:
                            candidates.append((alias.alias_name.lower(), p))
                    
                    for p in all_tenant_products:
                        if "MOCK" not in p.sku_id:
                            # Combine brand + sku_id + category into one searchable string
                            searchable = " ".join(filter(None, [
                                p.brand or "",
                                p.sku_id or "",
                                p.category or ""
                            ])).lower()
                            candidates.append((searchable, p))

                    for candidate_text, candidate_product in candidates:
                        matching_words = [w for w in token_words if w in candidate_text]
                        
                        if not matching_words:
                            continue
                        
                        # Require at least 2 matching words for multi-word tokens
                        # For single-word tokens, 1 match is sufficient
                        min_required = 2 if len(token_words) >= 2 else 1
                        if len(matching_words) < min_required:
                            continue
                        
                        # Score = ratio of token words matched (0.0 to 1.0)
                        score = len(matching_words) / len(token_words)
                        
                        if score > best_score:
                            best_score = score
                            best_product = candidate_product

                    # Accept match only if majority of token words matched (>=0.5)
                    if best_score >= 0.5 and best_product:
                        product = best_product
                        matched_in_phase2 = True
                        logger.info(
                            "IngestionService: Phase 2 word-overlap match '%s' -> '%s' (score %.2f)",
                            token, best_product.sku_id, best_score
                        )

            # Phase 3: Fuzzy Match Fallback
            if not product:
                try:
                    import rapidfuzz
                    # Build candidates if Phase 2 was skipped (empty token_words)
                    if not token_words:
                        candidates = []
                        for alias in all_tenant_aliases:
                            p = next((prod for prod in all_tenant_products if prod.id == alias.product_id), None)
                            if not p:
                                p = db.query(Product).filter_by(id=alias.product_id).first()
                            if p and "MOCK" not in p.sku_id:
                                candidates.append((alias.alias_name, p))
                        for p in all_tenant_products:
                            if "MOCK" not in p.sku_id:
                                candidates.append((p.sku_id, p))

                    if candidates:
                        choices = [c[0] for c in candidates]
                        candidate_map = {c[0]: c[1] for c in candidates}
                        result = rapidfuzz.process.extractOne(token, choices)
                        if result:
                            matched_text, score, index = result
                            if score >= 82.0:
                                product = candidate_map[matched_text]
                                matched_in_phase3 = True
                                logger.info(
                                    "IngestionService: Phase 3 fuzzy match '%s' -> '%s' (score %.2f)",
                                    token, matched_text, score
                                )
                except Exception as fuzzy_exc:
                    logger.error("IngestionService: Fuzzy matching error: %s", str(fuzzy_exc))

            # Phase 4 — Self-learning alias registration
            if product and (matched_in_phase2 or matched_in_phase3):
                if token_lower not in [a.alias_name.lower() for a in 
                    db.query(ProductAlias).filter(
                        ProductAlias.tenant_id == tenant_id,
                        ProductAlias.product_id == product.id
                    ).all()]:
                    try:
                        db.add(ProductAlias(
                            id=uuid.uuid4(),
                            tenant_id=tenant_id,
                            product_id=product.id,
                            alias_name=token_lower
                        ))
                        db.flush()
                        logger.info(
                            "IngestionService: Self-learned alias '%s' -> SKU %s",
                            token_lower, product.sku_id
                        )
                    except Exception as alias_err:
                        logger.warning("IngestionService: Alias self-learning failed: %s", str(alias_err))
                        db.rollback()

            if product:
                logger.info("IngestionService: Product matched: %s -> SKU %s", token, product.sku_id)
                parsed_items.append({
                    "product_name": token,
                    "sku_code": product.sku_id,
                    "sku_id": product.sku_id,
                    "brand": product.brand,
                    "category": product.category,
                    "pack_size": product.pack_size,
                    "qty": qty,
                    "wholesale_rate": float(product.base_price),
                    "rate": float(product.base_price),
                    "product_id": product.id,
                    "unmatched_raw_text": None
                })
            else:
                has_unmatched = True
                logger.warning("IngestionService: Product unmatched: %s", token)
                parsed_items.append({
                    "product_name": f"Unmatched: {token}",
                    "sku_code": "UNMATCHED_SKU",
                    "sku_id": "UNMATCHED_SKU",
                    "brand": token,
                    "category": "Grocery",
                    "pack_size": "1 unit",
                    "qty": qty,
                    "wholesale_rate": 0.0,
                    "rate": 0.0,
                    "product_id": None,
                    "unmatched_raw_text": token
                })

        # ── DATABASE COMMIT: Order & Line Item Creation ───────────────────────────
        generated_order_id = f"ORD-2506-{uuid.uuid4().hex[:4].upper()}"
        new_order = Order(
            id=uuid.uuid4(),
            tenant_id=tenant_id,
            internal_order_id=generated_order_id,
            source="WhatsApp",
            customer_id=customer.id,
            invoice_type=parsed_order.extracted_invoice_preference,
            created_at=datetime.utcnow(),
            raw_source_text=message_text
        )
        new_order.status = "pending_review" if has_unmatched else "Draft"
        db.add(new_order)
        db.flush()

        # Write line item child records to DB
        for item in parsed_items:
            db.add(OrderLineItem(
                id=uuid.uuid4(),
                tenant_id=tenant_id,
                order_id=new_order.id,
                product_id=item["product_id"],
                quantity=item["qty"],
                unit_price=item["rate"],
                unmatched_raw_text=item["unmatched_raw_text"]
            ))

        order_status = "pending_review" if has_unmatched else "Draft"
        new_order.status = order_status

        # Record state transition in ledger
        db.add(OrderStateLedger(
            tenant_id=tenant_id,
            order_id=new_order.id,
            from_status=None,
            to_status=order_status,
            updated_by="system_whatsapp_agent"
        ))

        db.commit()
        db.refresh(new_order)

        # Fire order_received notification (non-blocking)
        try:
            # Eagerly load relationships so they are in-memory before background task starts
            for item in new_order.line_items:
                if item.product:
                    _ = item.product.brand

            import asyncio
            from app.services.notification_service import NotificationService
            
            tenant_obj = db.query(DistributorTenant).filter_by(id=tenant_id).first()

            async def fire_notifications(tenant_val, customer_val, order_val):
                try:
                    notification_service = NotificationService(
                        evolution_base_url=os.getenv("EVOLUTION_API_URL", "http://34.158.60.42:8080"),
                        api_key=os.getenv("EVOLUTION_API_KEY", "distributorbotkey2026")
                    )
                    # Check if at least one line item is matched
                    has_matched = any(
                        item.product_id is not None and float(item.unit_price) > 0
                        for item in order_val.line_items
                    )
                    if not has_matched:
                        logger.info(f"Notification skipped: all items unmatched for order {order_val.id}")
                    else:
                        await notification_service.notify(
                            event="order_received",
                            tenant=tenant_val,
                            customer=customer_val,
                            order=order_val,
                            db=db
                        )
                    # Also fire distributor alert
                    if tenant_val.whatsapp_order_phone:
                        await notification_service.notify(
                            event="new_order_alert_to_distributor",
                            tenant=tenant_val,
                            customer=customer_val,  
                            order=order_val,
                            db=db,
                            override_to_phone=tenant_val.whatsapp_order_phone  # send to distributor's own number
                        )
                        # Order needs review: at least one line item couldn't be auto-matched.
                        # Fire a distinct, more urgent alert so this order isn't silently missed —
                        # per PRODUCT_STATUS.md this is the biggest week-1 churn risk.
                        has_unmatched_items = any(item.product_id is None for item in order_val.line_items)
                        if has_unmatched_items:
                            await notification_service.notify(
                                event="order_needs_review_alert",
                                tenant=tenant_val,
                                customer=customer_val,
                                order=order_val,
                                db=db,
                                override_to_phone=tenant_val.whatsapp_order_phone
                            )
                    else:
                        logger.warning("Distributor alert skipped: whatsapp_order_phone not configured for tenant %s", str(tenant_val.id))
                except Exception as inner_ex:
                    logger.warning("Notification fire failed silently: %s", str(inner_ex))

            try:
                loop = asyncio.get_running_loop()
            except RuntimeError:
                loop = None

            if loop and loop.is_running():
                loop.create_task(fire_notifications(tenant_obj, customer, new_order))
            else:
                asyncio.run(fire_notifications(tenant_obj, customer, new_order))
        except Exception as e:
            logger.warning("Notification fire setup failed silently: %s", str(e))

        logger.info(
            "IngestionService: Success! Order %s created totaling %s",
            generated_order_id,
            sum(item["qty"] * item["rate"] for item in parsed_items)
        )

        return {
            "status": "success",
            "order_id": generated_order_id,
            "job_id": str(uuid.uuid4()),
            "successful_rows": 1,
            "failed_rows": 0,
            "message": "Order captured successfully but requires manual assignment." if has_unmatched else "Ingestion completed successfully!",
            "error_message": None
        }
